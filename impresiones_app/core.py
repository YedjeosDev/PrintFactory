from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import Paragraph, Table, TableStyle


DEFAULT_SHEET = "INFO"
DEFAULT_ALIADO_COMERCIAL = "INMEL"
DEFAULT_TITLE = "FORMATO ENTREGA FACTURACIÓN ESPECIAL"
MAX_ROWS_PER_PAGE = 30

LEGAL_TEXT = (
    "A CARIBEMAR DE LA COSTA S.A.S. E.S.P., le son aplicables las normas que regulan el "
    "saneamiento contable en el sector público, por tanto, nos permitimos recordar la obligacion "
    "que nos asiste de reportar a la Contaduría General de la Nación, semestralmente el Boletín "
    "de Deudores Morosos del Estado - BDME - , y trimestralmente el informe de Recíprocas "
    "(Entidades oficiales), los cuales están reglamentados en la Resolución 706 de 2016 y en el "
    "parágrafo 3° del artículo 4° de la Ley 716 de 2001, modificado por el Artículo 2° de la Ley "
    "901 de 2004. Asi las cosas, previendo las consecuencias de quedar incluido en uno o ambos "
    "reportes, agradecemos a usted el pago oportuno y asi cumplir con las obligaciones de Ley.\n"
    "Nota: La deuda de las facturas relacionadas, no incluye deuda reclamada, en concordato, "
    "financiada, ni interés por mora. Estos se calculan diariamente y se liquidan el día que "
    "ingresa el pago efectivo de las facturas vencidas, considerando la tasa de interés vigente."
)

VISIBLE_COLUMNS = [
    "NIC",
    "SIMB VAR",
    "IMPORTE",
    "SALDO VENCIDO",
    "DIRECCION DE ENTREGA REAL",
    "SEGMENTO EJECUTIVO",
    "DESCRIPCIÓN DE SUMINISTRO",
    "TITULAR PAGO REAL",
]

REQUIRED_COLUMN_ALIASES = {
    "NIC": ["NIC"],
    "SIMB VAR": ["SIMB VAR"],
    "IMPORTE": ["IMPORTE"],
    "SALDO VENCIDO": ["SALDO VENCIDO"],
    "DIRECCION DE ENTREGA REAL": ["DIRECCION DE ENTREGA REAL"],
    "SEGMENTO EJECUTIVO": ["SEGMENTO EJECUTIVO"],
    "DESCRIPCIÓN DE SUMINISTRO": ["DESCRIPCIÓN DE SUMINISTRO", "DESCRIPCION DE SUMINISTRO"],
    "TITULAR PAGO REAL": ["TITULAR PAGO REAL", "TITULAR DE PAGO"],
}


@dataclass(frozen=True)
class SupplyRecord:
    nic: str
    simb_var: str
    importe: float
    saldo_vencido: float
    direccion_entrega: str
    segmento: str
    descripcion_suministro: str
    titular_pago: str

    @property
    def deuda(self) -> float:
        return self.saldo_vencido - self.importe

    def visible_row(self) -> list[str]:
        return [
            self.nic,
            self.simb_var,
            format_money(self.importe, with_symbol=False),
            format_money(self.saldo_vencido, with_symbol=False),
            self.direccion_entrega,
            self.segmento,
            self.descripcion_suministro,
            self.titular_pago,
        ]


@dataclass(frozen=True)
class PdfGroup:
    direccion_entrega: str
    segmento: str
    titular_pago: str
    records: list[SupplyRecord]

    @property
    def page_count(self) -> int:
        return max(1, math.ceil(len(self.records) / MAX_ROWS_PER_PAGE))


def normalize_header(value: object) -> str:
    text = str(value or "").strip().upper()
    text = re.sub(r"\s+", " ", text)
    return text


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value)).strip()


def parse_number(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    text = text.replace("$", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def format_money(value: float, with_symbol: bool = True) -> str:
    rounded = int(round(value or 0))
    formatted = f"{rounded:,}".replace(",", ".")
    return f"$ {formatted}" if with_symbol else formatted


def safe_filename(text: str, max_len: int = 120) -> str:
    safe = re.sub(r"[^\w\s.-]", "", text, flags=re.UNICODE)
    safe = re.sub(r"\s+", " ", safe).strip().replace(" ", "_")
    return (safe[:max_len] or "formato").strip("._")


def _build_header_map(headers: Iterable[object]) -> dict[str, int]:
    normalized = {normalize_header(header): index for index, header in enumerate(headers)}
    resolved: dict[str, int] = {}
    missing: list[str] = []

    for canonical, aliases in REQUIRED_COLUMN_ALIASES.items():
        found = None
        for alias in aliases:
            found = normalized.get(normalize_header(alias))
            if found is not None:
                break
        if found is None:
            missing.append(canonical)
        else:
            resolved[canonical] = found

    if missing:
        raise ValueError("Faltan columnas obligatorias: " + ", ".join(missing))
    return resolved


def load_excel_records(path: str | Path, sheet_name: str = DEFAULT_SHEET) -> list[SupplyRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}'. Hojas disponibles: {', '.join(workbook.sheetnames)}")

    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise ValueError("La hoja seleccionada está vacía.") from exc

    columns = _build_header_map(headers)
    records: list[SupplyRecord] = []

    for row in rows:
        if not row or all(value is None or str(value).strip() == "" for value in row):
            continue

        def get(name: str) -> object:
            index = columns[name]
            return row[index] if index < len(row) else None

        nic = clean_text(get("NIC"))
        direccion = clean_text(get("DIRECCION DE ENTREGA REAL"))
        segmento = clean_text(get("SEGMENTO EJECUTIVO"))

        if not nic or not direccion or not segmento:
            continue

        records.append(
            SupplyRecord(
                nic=nic,
                simb_var=clean_text(get("SIMB VAR")),
                importe=parse_number(get("IMPORTE")),
                saldo_vencido=parse_number(get("SALDO VENCIDO")),
                direccion_entrega=direccion,
                segmento=segmento,
                descripcion_suministro=clean_text(get("DESCRIPCIÓN DE SUMINISTRO")),
                titular_pago=clean_text(get("TITULAR PAGO REAL")),
            )
        )

    return records


def group_records(records: Iterable[SupplyRecord]) -> list[PdfGroup]:
    grouped: dict[tuple[str, str, str], list[SupplyRecord]] = {}
    for record in records:
        # El titular se incluye en la llave para no mezclar encabezados si un archivo trae datos inconsistentes.
        key = (record.direccion_entrega, record.segmento, record.titular_pago)
        grouped.setdefault(key, []).append(record)

    result = [
        PdfGroup(direccion, segmento, titular, items)
        for (direccion, segmento, titular), items in grouped.items()
    ]
    return sorted(result, key=lambda group: (group.direccion_entrega, group.segmento, group.titular_pago))


def generate_pdfs(
    records: Iterable[SupplyRecord],
    output_dir: str | Path,
    logo_path: str | Path | None = None,
    aliado_comercial: str = DEFAULT_ALIADO_COMERCIAL,
) -> list[Path]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    logo = Path(logo_path) if logo_path else None

    generated: list[Path] = []
    for index, group in enumerate(group_records(records), start=1):
        filename = safe_filename(f"{index:03d}_{group.titular_pago}_{group.segmento}_{group.direccion_entrega}") + ".pdf"
        pdf_path = output / filename
        _draw_group_pdf(pdf_path, group, logo, aliado_comercial)
        generated.append(pdf_path)

    return generated


def _draw_group_pdf(pdf_path: Path, group: PdfGroup, logo_path: Path | None, aliado_comercial: str) -> None:
    canvas = Canvas(str(pdf_path), pagesize=letter)
    total_pages = group.page_count

    for page_index in range(total_pages):
        start = page_index * MAX_ROWS_PER_PAGE
        end = start + MAX_ROWS_PER_PAGE
        page_records = group.records[start:end]
        _draw_page(
            canvas=canvas,
            group=group,
            page_records=page_records,
            absolute_start=start,
            page_number=page_index + 1,
            total_pages=total_pages,
            logo_path=logo_path,
            aliado_comercial=aliado_comercial,
        )
        canvas.showPage()

    canvas.save()


def _draw_page(
    canvas: Canvas,
    group: PdfGroup,
    page_records: list[SupplyRecord],
    absolute_start: int,
    page_number: int,
    total_pages: int,
    logo_path: Path | None,
    aliado_comercial: str,
) -> None:
    width, height = letter
    left = 14 * mm
    right = width - 14 * mm
    top = height - 13 * mm

    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        "small",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=5.2,
        leading=5.45,
        alignment=TA_LEFT,
        splitLongWords=False,
    )
    center = ParagraphStyle("small_center", parent=normal, alignment=TA_CENTER)
    right_style = ParagraphStyle("small_right", parent=normal, alignment=TA_RIGHT)
    bold_center = ParagraphStyle(
        "small_bold_center",
        parent=center,
        fontName="Helvetica-Bold",
        fontSize=5.6,
        leading=6.0,
    )
    bold_right = ParagraphStyle(
        "small_bold_right",
        parent=right_style,
        fontName="Helvetica-Bold",
        fontSize=5.6,
        leading=6.0,
    )
    legal_style = ParagraphStyle(
        "legal",
        parent=normal,
        fontSize=7.0,
        leading=8.5,
        alignment=TA_LEFT,
    )

    if logo_path and logo_path.exists():
        canvas.drawImage(str(logo_path), left, top - 18 * mm, width=31 * mm, height=18 * mm, preserveAspectRatio=True, mask="auto")

    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawCentredString(width / 2, top - 7 * mm, DEFAULT_TITLE)
    if total_pages > 1:
        canvas.setFont("Helvetica", 7)
        canvas.drawRightString(right, top - 7 * mm, f"Página {page_number} de {total_pages}")

    canvas.setFont("Helvetica-Bold", 7.3)
    header_y = top - 21 * mm
    label_x = left
    value_x = left + 36 * mm
    second_label_x = left + 120 * mm
    second_value_x = left + 155 * mm
    line_gap = 7.2 * mm

    def header_line(y: float, label: str, value: str, x_label: float = label_x, x_value: float = value_x) -> None:
        canvas.drawString(x_label, y, label)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(x_value, y, value[:62])
        canvas.setFont("Helvetica-Bold", 7.3)

    header_line(header_y, "ALIADO COMERCIAL:", aliado_comercial)
    header_line(header_y, "REPARTIDOR ASIGNADO:", "", second_label_x, second_value_x)
    header_line(header_y - line_gap, "TITULAR DE PAGO:", group.titular_pago)
    header_line(header_y - 2 * line_gap, "DIRECCION:", group.direccion_entrega)
    header_line(header_y - 3 * line_gap, "SEGMENTO / CUENTA:", group.segmento)

    table_top = header_y - 4.5 * line_gap
    row_height = 10.8
    col_widths = [24, 52, 178, 62, 72, 72, 72]

    data = [[
        Paragraph("No.", bold_center),
        Paragraph("NIC", bold_center),
        Paragraph("DESCRIPCION DE SUMINISTRO", bold_center),
        Paragraph("SIMB VAR", bold_center),
        Paragraph("IMPORTE MES", bold_center),
        Paragraph("DEUDA", bold_center),
        Paragraph("TOTAL", bold_center),
    ]]

    for offset, record in enumerate(page_records):
        data.append([
            Paragraph(str(absolute_start + offset + 1), center),
            Paragraph(record.nic, center),
            Paragraph(record.descripcion_suministro, normal),
            Paragraph(record.simb_var, center),
            Paragraph(format_money(record.importe), right_style),
            Paragraph(format_money(record.deuda), right_style),
            Paragraph(format_money(record.saldo_vencido), right_style),
        ])

    empty_rows = MAX_ROWS_PER_PAGE - len(page_records)
    for _ in range(empty_rows):
        data.append(["", "", "", "", "", "", ""])

    page_importe = sum(record.importe for record in page_records)
    page_deuda = sum(record.deuda for record in page_records)
    page_total = sum(record.saldo_vencido for record in page_records)

    label = "TOTAL"
    if total_pages > 1:
        label = "TOTAL PÁGINA"

    data.append([
        Paragraph(label, bold_right),
        "",
        "",
        "",
        Paragraph(format_money(page_importe), bold_right),
        Paragraph(format_money(page_deuda), bold_right),
        Paragraph(format_money(page_total), bold_right),
    ])

    table = Table(data, colWidths=col_widths, rowHeights=[12] + [row_height] * MAX_ROWS_PER_PAGE + [12])
    style_commands = [
        ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.1),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ("SPAN", (0, -1), (3, -1)),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F3F4F6")),
    ]
    for row_index in range(1, MAX_ROWS_PER_PAGE + 1):
        if row_index % 2 == 0:
            style_commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F7FBFF")))
    table.setStyle(TableStyle(style_commands))
    table.wrapOn(canvas, right - left, height)
    table.drawOn(canvas, left, table_top - table._height)

    legal_top = table_top - table._height - 5 * mm
    legal = Paragraph(LEGAL_TEXT.replace("\n", "<br/>"), legal_style)
    legal_width = right - left
    legal_height = 34 * mm
    legal.wrapOn(canvas, legal_width, legal_height)
    legal.drawOn(canvas, left, legal_top - legal_height + 8 * mm)

    footer_y = 18 * mm
    canvas.setFont("Helvetica-Bold", 8)
    canvas.drawString(left, footer_y, "FIRMA RECIBIDO:")
    canvas.line(left + 31 * mm, footer_y - 1, left + 105 * mm, footer_y - 1)
    canvas.drawString(left + 118 * mm, footer_y, "FECHA RECIBIDO:")
    canvas.line(left + 149 * mm, footer_y - 1, right, footer_y - 1)
